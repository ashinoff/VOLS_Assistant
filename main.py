import telebot
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton
import pandas as pd
import requests
from io import StringIO, BytesIO
from dotenv import load_dotenv
import os
from datetime import datetime
import openpyxl
from flask import Flask, request, abort
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

load_dotenv()
bot = telebot.TeleBot(os.getenv('TELEGRAM_TOKEN'))
app = Flask(__name__)

ZONES_URL = os.getenv('ZONES_CSV_URL')

rk_branches = {
    "Юго-Западные ЭС": {"key": "YUGO_ZAPADNYE_ES", "filial": "Юго-Западный"},
    "Усть-Лабинские ЭС": {"key": "UST_LABINSKIE_ES", "filial": "Усть-Лабинский"},
    "Тимашевские ЭС": {"key": "TIMASHEVSKIE_ES", "filial": "Тимашевский"},
    "Тихорецкие ЭС": {"key": "TIHORETSKIE_ES", "filial": "Тихорецкий"},
    "Сочинские ЭС": {"key": "SOCHINSKIE_ES", "filial": "Сочинский"},
    "Славянские ЭС": {"key": "SLAVYANSKIE_ES", "filial": "Славянский"},
    "Ленинградские ЭС": {"key": "LENINGRADSKIE_ES", "filial": "Ленинградский"},
    "Лабинские ЭС": {"key": "LABINSKIE_ES", "filial": "Лабинский"},
    "Краснодарские ЭС": {"key": "KRASNODARSKIE_ES", "filial": "Краснодарский"},
    "Армавирские ЭС": {"key": "ARMAVIRSKIE_ES", "filial": "Армавирский"},
    "Адыгейские ЭС": {"key": "ADYGEYSKIE_ES", "filial": "Адыгейский"},
}

ug_branches = {
    "Юго-Западные ЭС": {"key": "YUGO_ZAPADNYE_ES", "filial": "Юго-Западный"},
    "Центральные ЭС": {"key": "TSENTRALNYE_ES", "filial": "Центральный"},
    "Западные ЭС": {"key": "ZAPADNYE_ES", "filial": "Западный"},
    "Восточные ЭС": {"key": "VOSTOCHNYE_ES", "filial": "Восточный"},
    "Южные ЭС": {"key": "YUZHNIE_ES", "filial": "Южный"},
    "Северо-Восточные ЭС": {"key": "SEVERO_VOSTOCHNYE_ES", "filial": "Северо-Восточный"},
    "Юго-Восточные ЭС": {"key": "YUGO_VOSTOCHNYE_ES", "filial": "Юго-Восточный"},
    "Северные ЭС": {"key": "SEVERNYE_ES", "filial": "Северный"},
}

rk_notifications = []
ug_notifications = []
user_state = {}

def load_csv(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        df = pd.read_csv(StringIO(response.text), sep='\t', encoding='utf-8')
        logging.info(f"Loaded CSV from {url} with columns: {df.columns.tolist()}")
        return df
    except Exception as e:
        logging.error(f"Error loading CSV from {url}: {e}")
        return pd.DataFrame()

def get_user_zone(user_id):
    zones_df = load_csv(ZONES_URL)
    if zones_df.empty:
        return None
    zones_df['Telegram ID'] = pd.to_numeric(zones_df['Telegram ID'], errors='coerce')
    user_row = zones_df[zones_df['Telegram ID'] == user_id]
    return user_row.iloc[0] if not user_row.empty else None

def build_main_keyboard(vis, fil, res):
    markup = ReplyKeyboardMarkup(resize_keyboard=True)
    if vis == 'All':
        markup.add("РОССЕТИ ЮГ", "РОССЕТИ КУБАНЬ", "ТЕЛЕФОНЫ КОНТРАГЕНТОВ", "ОТЧЕТЫ", "СПРАВКА")
    elif vis == 'RK':
        branches_dict = rk_branches
        if fil == 'All':
            markup.add("РОССЕТИ КУБАНЬ", "ТЕЛЕФОНЫ КОНТРАГЕНТОВ", "ОТЧЕТЫ", "СПРАВКА")
        else:
            branch = next(k for k, v in branches_dict.items() if v['filial'] == fil)
            markup.add(branch, "ТЕЛЕФОНЫ КОНТРАГЕНТОВ")
            if res == 'All': markup.add("ОТЧЕТЫ")
            markup.add("СПРАВКА")
    elif vis == 'UG':
        branches_dict = ug_branches
        if fil == 'All':
            markup.add("РОССЕТИ ЮГ", "ТЕЛЕФОНЫ КОНТРАГЕНТОВ", "ОТЧЕТЫ", "СПРАВКА")
        else:
            branch = next(k for k, v in branches_dict.items() if v['filial'] == fil)
            markup.add(branch, "ТЕЛЕФОНЫ КОНТРАГЕНТОВ")
            if res == 'All': markup.add("ОТЧЕТЫ")
            markup.add("СПРАВКА")
    return markup

@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.from_user.id
    logging.info(f"User {user_id} sent /start")
    try:
        zone = get_user_zone(user_id)
        if zone is None:
            logging.info(f"No access for user {user_id}")
            bot.send_message(message.chat.id, "Нет доступа.")
            return
        bot.send_message(message.chat.id, "Добро пожаловать!", reply_markup=build_main_keyboard(zone['Видимость'], zone['Филиал'], zone['РЭС']))
    except Exception as e:
        logging.error(f"Error in start for user {user_id}: {e}")
        bot.send_message(message.chat.id, "Ошибка запуска.")

@bot.message_handler(func=lambda m: True)
def handle_text(message):
    text = message.text
    user_id = message.from_user.id
    zone = get_user_zone(user_id)
    if zone is None: return
    vis, fil, res = zone['Видимость'], zone['Филиал'], zone['РЭС']
    state = user_state.get(user_id, {})
    if 'stage' in state:
        if state['stage'] == 'search_tp_input':
            norm_input = ''.join(c.lower() for c in text if c.isalnum())
            group = state['group']
            branch = state['branch']
            branches_dict = rk_branches if group == 'RK' else ug_branches
            info = branches_dict[branch]
            url = os.getenv(info['key'] + '_URL_' + group)
            df = load_csv(url)
            if df.empty:
                bot.send_message(message.chat.id, "Ошибка данных.")
                del state['stage']
                return
            df['norm_tp'] = df['Наименование ТП'].apply(lambda x: ''.join(c.lower() for c in str(x) if c.isalnum()))
            matching = df[df['norm_tp'].str.contains(norm_input)]
            tps = matching['Наименование ТП'].unique()
            if len(tps) == 0:
                bot.send_message(message.chat.id, "Не найдено.")
            else:
                markup = InlineKeyboardMarkup()
                for tp in tps:
                    markup.add(InlineKeyboardButton(tp, callback_data=f"search_tp_{tp}"))
                markup.add(InlineKeyboardButton("Назад", callback_data="back"))
                bot.send_message(message.chat.id, "Выберите ТП", reply_markup=markup)
            del state['stage']
            user_state[user_id] = state
            return
        elif state['stage'] == 'notify_tp_input':
            norm_input = ''.join(c.lower() for c in text if c.isalnum())
            group = state['group']
            branch = state['branch']
            branches_dict = rk_branches if group == 'RK' else ug_branches
            info = branches_dict[branch]
            url = os.getenv(info['key'] + '_URL_' + group + '_SP')
            df = load_csv(url)
            if df.empty:
                bot.send_message(message.chat.id, "Ошибка данных.")
                del state['stage']
                return
            df['norm_tp'] = df['Наименование ТП'].apply(lambda x: ''.join(c.lower() for c in str(x) if c.isalnum()))
            matching = df[df['norm_tp'].str.contains(norm_input)]
            tps = matching['Наименование ТП'].unique()
            if len(tps) == 0:
                bot.send_message(message.chat.id, "Не найдено.")
            else:
                markup = InlineKeyboardMarkup()
                for tp in tps:
                    markup.add(InlineKeyboardButton(tp, callback_data=f"notify_tp_{tp}"))
                markup.add(InlineKeyboardButton("Назад", callback_data="back"))
                bot.send_message(message.chat.id, "Выберите ТП", reply_markup=markup)
            del state['stage']
            user_state[user_id] = state
            return
    if text in ["РОССЕТИ КУБАНЬ", "РОССЕТИ ЮГ"]:
        group = 'RK' if text == "РОССЕТИ КУБАНЬ" else 'UG'
        branches = rk_branches if group == "RK" else ug_branches
        if (vis == 'All' or vis == group) and fil == 'All':
            markup = ReplyKeyboardMarkup(resize_keyboard=True)
            for b in branches:
                markup.add(b)
            markup.add("Назад")
            bot.send_message(message.chat.id, "Выберите филиал", reply_markup=markup)
            user_state[user_id] = {'group': group}
    elif text in rk_branches or text in ug_branches:
        if user_id in user_state and 'group' in user_state[user_id]:
            group = user_state[user_id]['group']
        else:
            group = vis
        branches_dict = rk_branches if group == 'RK' else ug_branches
        if text not in branches_dict:
            return
        info = branches_dict[text]
        if (vis == 'All' or vis == group) and (fil == 'All' or fil == info['filial']):
            markup = ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add("Поиск по ТП", "Отправить уведомление", "Справка", "Назад")
            bot.send_message(message.chat.id, "Меню филиала", reply_markup=markup)
            user_state[user_id] = {'group': group, 'branch': text}
    elif text == "Поиск по ТП":
        if user_id in user_state and 'branch' in user_state[user_id]:
            bot.send_message(message.chat.id, "Введите наименование ТП")
            user_state[user_id]['stage'] = 'search_tp_input'
    elif text == "Отправить уведомление":
        if user_id in user_state and 'branch' in user_state[user_id]:
            bot.send_message(message.chat.id, "Введите наименование ТП")
            user_state[user_id]['stage'] = 'notify_tp_input'
    elif text == "ОТЧЕТЫ":
        markup = ReplyKeyboardMarkup(resize_keyboard=True)
        if vis == 'All':
            markup.add("Уведомления РОССЕТИ КУБАНЬ", "Уведомления РОССЕТИ ЮГ")
        elif vis == 'RK':
            markup.add("Уведомления РОССЕТИ КУБАНЬ")
        elif vis == 'UG':
            markup.add("Уведомления РОССЕТИ ЮГ")
        markup.add("Назад")
        bot.send_message(message.chat.id, "Выберите отчет", reply_markup=markup)
    elif text in ["Уведомления РОССЕТИ КУБАНЬ", "Уведомления РОССЕТИ ЮГ"]:
        group = 'RK' if "КУБАНЬ" in text else 'UG'
        notifs = rk_notifications if group == 'RK' else ug_notifications
        if not notifs:
            bot.send_message(message.chat.id, "Нет уведомлений.")
            return
        df = pd.DataFrame(notifs)
        if fil != 'All':
            df = df[df['ФИЛИАЛ'] == fil]
        output = BytesIO()
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        pink = openpyxl.styles.PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
        for cell in ws[1]:
            cell.fill = pink
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        bot.send_document(message.chat.id, output, visible_file_name='report.xlsx')
    elif text == "СПРАВКА":
        markup = ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add("Форма доп соглашения", "Форма претензии", "Назад")
        bot.send_message(message.chat.id, "Выберите", reply_markup=markup)
    elif text == "Форма доп соглашения":
        url = os.getenv('DOP_SOG_URL')
        if url:
            bot.send_document(message.chat.id, url)
        else:
            bot.send_message(message.chat.id, "Ссылка не найдена.")
    elif text == "Форма претензии":
        url = os.getenv('PRETENSII_URL')
        if url:
            bot.send_document(message.chat.id, url)
        else:
            bot.send_message(message.chat.id, "Ссылка не найдена.")
    elif text == "ТЕЛЕФОНЫ КОНТРАГЕНТОВ":
        bot.send_message(message.chat.id, "В разработке.")
    elif text == "Назад":
        bot.send_message(message.chat.id, "Назад", reply_markup=build_main_keyboard(vis, fil, res))
        if user_id in user_state: del user_state[user_id]

@bot.callback_query_handler(func=lambda call: True)
def handle_callback(call):
    data = call.data
    user_id = call.from_user.id
    state = user_state.get(user_id, {})
    group = state.get('group')
    branch = state.get('branch')
    branches_dict = rk_branches if group == 'RK' else ug_branches
    info = branches_dict.get(branch, {})
    if data.startswith('search_tp_'):
        tp = data[10:]
        url = os.getenv(info['key'] + '_URL_' + group)
        df = load_csv(url)
        results = df[df['Наименование ТП'] == tp]
        if results.empty: return
        num = len(results)
        msg = f"{results.iloc[0]['РЭС']} РЭС, на {tp} найдено {num} ВОЛС с договором аренды.\n"
        for _, row in results.iterrows():
            msg += f"* ВЛ: {row['Наименование ВЛ']}\nОпоры: {row['Опоры']}, Количество опор: {row['Количество опор']}\nКонтрагент: {row['Наименование Провайдера']}\n\n"
        bot.send_message(call.message.chat.id, msg)
    elif data.startswith('notify_tp_'):
        tp = data[10:]
        user_state[user_id]['selected_tp'] = tp
        url = os.getenv(info['key'] + '_URL_' + group + '_SP')
        df = load_csv(url)
        vls = df[df['Наименование ТП'] == tp]['Наименование ВЛ'].unique()
        markup = InlineKeyboardMarkup()
        for vl in vls:
            markup.add(InlineKeyboardButton(vl, callback_data=f"notify_vl_{vl}"))
        markup.add(InlineKeyboardButton("Назад", callback_data="back"))
        bot.send_message(call.message.chat.id, "Выберите ВЛ", reply_markup=markup)
    elif data.startswith('notify_vl_'):
        vl = data[10:]
        user_state[user_id]['selected_vl'] = vl
        markup = ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(KeyboardButton("Отправить геоданные", request_location=True))
        bot.send_message(call.message.chat.id, "Отправьте геоданные", reply_markup=markup)

@bot.message_handler(content_types=['location'])
def handle_location(message):
    user_id = message.from_user.id
    state = user_state.get(user_id, {})
    if 'selected_vl' not in state: return
    lat, lon = message.location.latitude, message.location.longitude
    tp, vl = state['selected_tp'], state['selected_vl']
    group, branch = state['group'], state['branch']
    branches_dict = rk_branches if group == 'RK' else ug_branches
    info = branches_dict[branch]
    url_sp = os.getenv(info['key'] + '_URL_' + group + '_SP')
    df_sp = load_csv(url_sp)
    matching = df_sp[(df_sp['Наименование ТП'] == tp) & (df_sp['Наименование ВЛ'] == vl)]
    if matching.empty: return
    row = matching.iloc[0]
    filial_sp, res_sp = row['Филиал'], row['РЭС']
    zones_df = load_csv(ZONES_URL)
    resp_ids = zones_df[zones_df['Ответственный'].isin([filial_sp, res_sp])]['Telegram ID'].unique()
    if len(resp_ids) == 0:
        bot.send_message(message.chat.id, f"Ответственный не назначен на {res_sp} РЭС")
    else:
        sender_fio = get_user_zone(user_id)['ФИО']
        msg = f"{sender_fio} нашел бездоговорной ВОЛС на {tp}, {vl}."
        for rec_id in resp_ids:
            bot.send_message(rec_id, msg)
            bot.send_location(rec_id, lat, lon)
        bot.send_message(message.chat.id, f"Уведомление отправлено ответственному за {res_sp} РЭС")
        time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        coords = f"{lat},{lon}"
        receivers = zones_df[zones_df['Telegram ID'].isin(resp_ids)]
        for _, rec in receivers.iterrows():
            notif = {
                'ФИЛИАЛ': filial_sp, 'РЭС': res_sp, 'ФИО ОТПРАВИТЕЛЯ': sender_fio, 'id ОТПРАВИТЕЛЯ': user_id,
                'ФИО ПОЛУЧАТЕЛЯ': rec['ФИО'], 'id ПОЛУЧАТЕЛЯ': rec['Telegram ID'], 'ВРЕМЯ ДАТА': time_str, 'КООРДИНАТЫ': coords
            }
            (rk_notifications if group == 'RK' else ug_notifications).append(notif)
    del state['selected_tp']
    del state['selected_vl']

@app.route('/', methods=['GET', 'HEAD'])
def index():
    return 'OK'

@app.route('/', methods=['POST'])
def webhook():
    if request.headers.get('content-type') == 'application/json':
        json_string = request.get_data().decode('utf-8')
        update = telebot.types.Update.de_json(json_string)
        bot.process_new_updates([update])
        return ''
    else:
        abort(403)

if __name__ == '__main__':
    import time
    bot.remove_webhook()
    time.sleep(0.1)
    bot.set_webhook(url=os.getenv('WEBHOOK_URL'))
    app.run(host='0.0.0.0', port=int(os.getenv('PORT', 5000)))
